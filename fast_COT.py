import sys

import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup as bs
import requests
import openpyxl
from copy import copy
import datetime

import locale
from openpyxl.formula import Tokenizer

def trouver_date_dernier_mardi():
    dico={'Friday':-10,'Saturday':-4,'Sunday':-5,'Monday':-6,'Tuesday':-7,'Wednesday':-8,'Thursday':-9}
    now=datetime.datetime.now()
    jour=now.strftime("%A")
    td=datetime.timedelta(dico[jour])
    date_cot=now+td
    date=datetime.datetime(year=date_cot.year,month=date_cot.month,day=date_cot.day)
    return date

def main_COT(chemin_fichier):
    #cela sert à convertir les chiffre
    locale.setlocale(locale.LC_ALL, 'nl_NL')
    #URL pour chercher les informations
    url1 = "https://www.cftc.gov/dea/futures/deacmesf.htm"
    url2 = "https://www.cftc.gov/dea/futures/deanybtsf.htm"
    url3 = "https://www.cftc.gov/dea/futures/deacmxsf.htm"





    dico_actifs = [{"site": requests.get(url1),
                "liste_actifs": [{"name_site": "EURO FX - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "EUR"}
                    , {"name_site": "CANADIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "CAD"},
                                 {"name_site": "JAPANESE YEN - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "JPY"},
                                 {"name_site": "AUSTRALIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "AUD"},
                                 {"name_site": "SWISS FRANC - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "CHF"},
                                 {"name_site": "BITCOIN - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "BTC"},
                                 {"name_site": "ETHER CASH SETTLED - CHICAGO MERCANTILE EXCHANGE" , "name_sheet": "ETH"},
                                 {"name_site": "NZ DOLLAR - CHICAGO MERCANTILE EXCHANGE", "name_sheet": "NZD"},
                                 {"name_site": "BRITISH POUND - CHICAGO MERCANTILE EXCHANGE ", "name_sheet": "GBP"},
                                 {"name_site": "P 500 Consolidated - CHICAGO MERCANTILE EXCHANGE",
                                  "name_sheet": "S&P500"},{"name_site": "MEXICAN PESO - CHICAGO MERCANTILE EXCHANGE ",
                                  "name_sheet": "MXN"},{"name_site": "SO AFRICAN RAND - CHICAGO MERCANTILE EXCHANGE",
                                  "name_sheet": "ZAR"},
                                {"name_site": "BRAZILIAN REAL - CHICAGO MERCANTILE EXCHANGE",
                                  "name_sheet": "BRL"}
                                 ]},
               {"site": requests.get(url2),
                "liste_actifs": [{"name_site": 'USD INDEX - ICE FUTURES U.S.', "name_sheet": "USD"}]},
               {"site": requests.get(url3),
                "liste_actifs": [{"name_site": "SILVER - COMMODITY EXCHANGE INC.", "name_sheet": "SILVER"}
                    , {"name_site": "GOLD - COMMODITY EXCHANGE INC.", "name_sheet": "GOLD"}]}]

    résultat = {}

    for i in dico_actifs:
        soup = bs(i['site'].text, 'html.parser')
        text = soup.find_all("pre")
        liste_lignes = str(text).splitlines()
        for paire in i["liste_actifs"]:
            for z in range(len(liste_lignes)):
                if liste_lignes[z].find(paire['name_site']) != -1:
                    if liste_lignes[z].find("MICRO") != -1:
                        continue
                    else:
                        l = liste_lignes[z + 9].split()
                        if l[0].find(",")==-1:
                           l[0]=f"0,{l[0]}"
                        if l[1].find(",")==-1:
                            l[1]=f"0,{l[1]}"
                        résultat[paire['name_sheet']] = {"long": locale.atof(l[0]), "short": locale.atof(l[1])}
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    wb = openpyxl.load_workbook(chemin_fichier)
    dico_sheet = {i.title: i for i in wb}

    date = trouver_date_dernier_mardi()
    # gestion du style des cellules des cellules insérer
    cellule_style = dico_sheet['EUR']["B2"]
    font = cellule_style.font
    border = cellule_style.border
    fill = cellule_style.fill
    alignment = cellule_style.alignment
    decallage_début = "A2"

    for paire, données in résultat.items():
        collonnes_values = {'A': date, 'B': données['long'], 'C': données['short'], 'D': 0, 'E': 0, 'F': 0, 'G': 0}
        df = pd.read_excel(chemin_fichier, engine="openpyxl", sheet_name=f"{paire}")
        # donne la longueur du dataframme pour le décaller ensuite
        count = 0
        ws = dico_sheet[paire]
        if ws['A1'].value==date:
            continue
        for z, value in enumerate(df['SHORT'].isna()):
            if not value:
                count += 1


        decallage_fin = f"G{count}"
        ws.move_range(f"{decallage_début}:{decallage_fin}", rows=1, translate=True)

        # création des différentes valeurs pour les cellules
        value_ancien_long = ws['B3'].value
        value_ancien_short = ws['C3'].value
        value_ancien_net_position=ws['F3'].value
        if isinstance(value_ancien_net_position,str):
           p=Tokenizer(value_ancien_net_position)
        cloture_long = round(value_ancien_long-collonnes_values['B'],3)
        cloture_short=round(value_ancien_short-collonnes_values['C'],3)
        net_position=collonnes_values['B']-collonnes_values['C']
        pourcent_variation=round(((net_position-float(value_ancien_net_position))/abs(float(value_ancien_net_position))),2)
        #update des values du dico de valeurs
        collonnes_values.update({"D":cloture_long,"E":cloture_short,"F":net_position,'G':pourcent_variation})
        # attribution des styles au différents cellules et valeurs
        for c, v in collonnes_values.items():
            #attribution des styles et valeurs
            cellule = ws[f"{c}2"]
            cellule.font = copy(font)
            cellule.border = copy(border)
            cellule.fill = copy(fill)
            cellule.alignment = copy(alignment)
            cellule.number_format = copy(ws[f"{c}3"].number_format)
            cellule.value = v


    wb.save(chemin_fichier)


#ca à l'air de marché faire d'autres tests notamment en effacant les 2 premieres données
# vérifier que le SNP et ETH marche bien et sinon essayer de detecter pourquoi ca marche pas;

#H-J 3

### faire le script pour ajouter les datas au deuxieme fichier excel