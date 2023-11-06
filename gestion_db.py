import datetime
import sqlite3
import pandas as pd
import openpyxl
#{'Date': '2023-09-03 22:45:00.0', 'Pays': 'Nouvelle-Zélande', 'Devise': 'NZD', 'nom_annonce': "Prix À L'Exportation En Glissement Trimestriel (Q2)", 'I
#mpact': 'FAIBLE', 'Précédent': '-6.9%', 'Concensus': '-4.2%'}
def insert_next_annonce(datas):
    conn=sqlite3.connect("db.db")
    cu=conn.cursor()
    for annonce in datas:
        liste_key=["nom_annonce","Concensus"]
        for key in liste_key:
            if key not in annonce:
                annonce[key]=""
        if "Heure" not in annonce:
            date,heure=annonce["Date"].split(" ")
            heure=datetime.datetime.strptime(heure,"%H:%M:%S.%f")
            heure=(heure+datetime.timedelta(hours=2)).strftime("%H:%M:%S.%f")
            annonce['Heure']=heure
            annonce['Date']=date


        try:
            cu.execute('''
                    INSERT INTO Annonce_semaine (Date,Heure, Pays, Devise, Nom_annonce, Impact, Précédent, Concensus)
                    VALUES (?,?, ?, ?, ?, ?, ?, ?)
                ''', (
            annonce['Date'],annonce["Heure"],annonce['Pays'], annonce['Devise'], annonce['nom_annonce'], annonce['Impact'], annonce['Précédent'],
            annonce['Concensus']))
        except:print('Fuck')
    conn.commit()
    conn.close()

def remove_data():
    conn=sqlite3.connect("db.db")
    cu=conn.cursor()
    cu.execute("DELETE FROM Annonce_semaine")
    conn.commit()
    conn.close()



def insert_week_post_annonce(datas,chemin_excel):
    df=pd.DataFrame(datas)
    dfs_annonces=df.groupby(["Devise"])
    # Charger le fichier Excel

    wb = openpyxl.load_workbook(chemin_excel)
    for devise,df_annonces_devise in dfs_annonces:
        if devise in wb.sheetnames:
            sheet = wb[devise]

            # Convertir le DataFrame en une structure que openpyxl peut insérer
            data = df_annonces_devise.values.tolist()

            # Insérer les données dans la feuille existante
            for row in data:
                sheet.append(row)
        else:
            #création d'une feuille et insertion des données
            new_sheet = wb.create_sheet(title=devise, index=0)
            new_sheet.append(df_annonces_devise.columns.tolist())
            # Convertir le DataFrame en une structure que openpyxl peut insérer
            data = df_annonces_devise.values.tolist()
            # Insérer les données dans la nouvelle feuille
            for row in data:
                new_sheet.append(row)
    wb.save(chemin_excel)











