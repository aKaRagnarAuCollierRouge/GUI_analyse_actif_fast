import os
import re

import pandas as pd
from datetime import datetime
import pandas as pd
import openpyxl
import sys
import openpyxl
import calendar
import locale
import datetime

locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')

def créer_nouvelle_feuille(chemin,nom_feuille,listes_entêtes:list):
    classeur=openpyxl.load_workbook(chemin)

    if nom_feuille not in classeur.sheetnames:
        feuille=classeur.create_sheet(title=nom_feuille)

        feuille.append(listes_entêtes)
    classeur.save(chemin)

#chemin d'un fichier txt "C:\\Users\Baptiste\\Desktop\\Backtesting\\Analyse sentiment forex\\Banque de donnés\\2023\\juillet\\18\\11_35.txt"
#OK ca marche et trie les données sous cette forme [{paire:...,short:...,long:...,date:...,heure:....},{...}]
def trie_fichier_txt(chemin_fichier):
    month_list=["janvier","février","mars","avril","mai","juin","juillet","août","septembre","octobre","novembre","décembre"]

    with open(chemin_fichier,"r") as fichier:
        fichier_lu=fichier.read()
    datas=[]
    liste_lignes=fichier_lu.splitlines()
    first_line=liste_lignes[0].replace("POSITIONNEMENT DUMB MONNEY |","")
    split_fisrt_line=first_line.split(" à ")
    heure=datetime.datetime.strptime(split_fisrt_line[1],"%H:%M:%S.%f").time()
    # Utilisation d'une expression régulière pour extraire la date
    jour_finish=False
    year=""
    month=""
    day=""
    for index,char in enumerate(split_fisrt_line[0]):
        if char.isdigit() and char!= " ":
            if not jour_finish:
                day+=char
                if  split_fisrt_line[0][index+1].isalpha() :
                    jour_finish=True
            else:
                year+=char
        elif char.isalpha() and char!=" ":
            month+=char

    date = datetime.datetime(day=int(day),month=month_list.index(month)+1,year=int(year)).date()

    for i in range(2,len(liste_lignes)):
        data=liste_lignes[i].replace("Long:", "").replace("Short:", "").replace('Lot_long:',"").replace("Lot_short:","")
        datas_arrangé=data.split()
        if len(datas_arrangé)<=3:
            datas_arrangé.append('0')
            datas_arrangé.append('0')
        datas.append({"Date":date,"Heure":heure,"Paire":datas_arrangé[0],"Short":int(datas_arrangé[2]),"Long":int(datas_arrangé[1]),"Lot_long":float(datas_arrangé[3]),"Lot_short":float(datas_arrangé[4])})

    return datas

def conversion_to_df(liste_dico):
    df = pd.DataFrame(liste_dico)
    division_paire=df.groupby("Paire")
    return division_paire
    #Permet de créer un ittérable qui renvoie les df séparé par les paires.
    #Les données sont bien triées  par paire --> Créer un tuple de la forme ('paire',df)




def trie_and_traitement_dossier_fichier_one_month(main_chemin,year,month):
    month_list=["janvier","février","mars","avril","mai","juin","juillet","août","septembre","octobre","novembre","décembre"]
    if isinstance(month,str):
        if month.isdigit():
                month = month_list[int(month)-1]
    elif isinstance(month,int):
        month=month_list[month-1]
    main_chemin+=f"\\{year}\\{month}"

    liste_dossier_jour=os.listdir(main_chemin)
    liste_datas=[]
    for dossier in liste_dossier_jour:
        c=main_chemin+f"\\{dossier}"
        liste_fichier_txt=os.listdir(c)
        for fichier in liste_fichier_txt:
            chemin_final=c+f"\\{fichier}"
            data=trie_fichier_txt(chemin_final)
            for i in data:
                liste_datas.append(i)
    return liste_datas

# OK donc à partir de là j'ai tout les fichier txt du mois qui on était mit en dictionnaire:
# il prenne la forme liste=[{"date":,paire:,heure:,short:,long:},{.....}]
            #PAS TRES SMART PARCE QUE JOUVRE LE FICHIER ET LE FERME A CHAQUE FOIS, IL FAUT QUE JE TRIE MES DONNES AUTREMENT
            #Creer un dataframme avec Date|Heure|actifs|short|long puis retrier par Actif en créant plusieurs DF, enfin trier par date et heure et écrire data dans excel

#après avoir séparrer les dataframmes avec ma fonction conversion to df
def écriture_fin_excel(chemin_excel,liste_paire):

    wb = openpyxl.load_workbook(chemin_excel)
    for tuple_paire in liste_paire:
        # Vérifiez si la feuille existe
        if tuple_paire[0] in wb.sheetnames:
            feuille = wb[tuple_paire[0]]
        else:
            feuille = wb.create_sheet(tuple_paire[0])
            entetes = ["Date", "Heure", "Paire", "Long", "Short","Lot_long","Lot_short"]
            feuille.append(entetes)
        tuple_paire_trie=tuple_paire[1].sort_values(by=['Date', 'Heure'])
        for _,ligne in tuple_paire_trie.iterrows():
            feuille.append(list(ligne))  # append permet d'ecrire apres les dernieres datas du fichier excel
    wb.save(chemin_excel)



main="C:\\Users\Baptiste\\Desktop\\Backtesting\\Analyse sentiment forex\\Banque de donnés"
chemin_excel="C:\\Users\\Baptiste\\Desktop\\Backtesting\\Analyse sentiment forex\\Sentiment_forex_analyse.xlsx"
def report_sentiment_to_excel_previous_month(chemin_excel=chemin_excel,main_chemin=main):
    now=datetime.datetime.now()
    month=now.month-1
    year=now.year
    #si changement d'année
    if month<1:
        month=12
        year=year-1
    liste_dico= trie_and_traitement_dossier_fichier_one_month(main_chemin, year, month)
    liste_paire = conversion_to_df(liste_dico)
    écriture_fin_excel(chemin_excel, liste_paire)




def report_sentiment_to_excel_month(year:int,month:int,chemin_excel=chemin_excel,main_chemin=main):
    liste_dico = trie_and_traitement_dossier_fichier_one_month(main_chemin, year, month)
    liste_paire = conversion_to_df(liste_dico)
    écriture_fin_excel(chemin_excel, liste_paire)




